#%%
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 11:28:37 2023

@author: kakshay
"""
import pandas as pd
import openpyxl
import os

#%%
# Specify the input .plt file and output CSV file paths
input_file = "test_6.plt"
excel_output_file = os.path.splitext(input_file)[0] + ".xlsx"
csv_output_file = os.path.splitext(input_file)[0] + ".csv"
# Initialize variables to store data
headers = ["Timestamp(s)"]
data = []

# Read the .plt file
with open(input_file, "r") as plt_file:
    parsed_data = plt_file.readlines()

header_data = [] 
for item in parsed_data:
    try:
        header_data.append(float(item))
    except ValueError:
        header_data.append(item)

index = None
# Parse the header information
for i, value in enumerate(header_data):
    if value == 0.0:
        index = i
        break
    if i % 2 == 0 and i >= 4:
        # Even-indexed lines are headers
        header = value.strip()
        headers.append(header)
                 
# Parse the data section
for value in parsed_data[index:]:
    columns = value.strip().split()  # Split columns based on spaces or tabs  
    data.append(columns)
#%%
#TO parse the data vertically with index as TImestamp 
# Create a DataFrame by transposing the data
df0 = pd.DataFrame(data).T

# Convert to numeric
df1 = df0.apply(pd.to_numeric, errors='coerce')

# Create a new DataFrame to store the result
df2 = pd.DataFrame()

#number of columns 
col_numbers = len(headers)

# Calculate the number of iterations needed
num_iterations = len(df1.columns) // col_numbers

# Loop through the iterations
for i in range(num_iterations):
    start_col = i * col_numbers
    end_col = min((i + 1) * col_numbers, len(df1.columns))
    chunk = df1.iloc[:, start_col:end_col]
    
    chunk.columns = range(col_numbers)
    
    # Concatenate the chunk with the result DataFrame
    df2 = pd.concat([df2, chunk], axis=0, ignore_index=True)

# Assign the specified headers to create df2
df2 = df2.rename(columns=dict(zip(df1.columns, headers)))

# Fill NaN values in Timestamp column with previous value
df2= df2.fillna(method='ffill')

# Export the DataFrame to a CSV file
df2.to_csv(csv_output_file, index=False)

# Write the DataFrame to an Excel file
df2.to_excel(excel_output_file, index=False)  # Set index=False to exclude the DataFrame index in the output

#%%
# Load the Excel file to apply the column heading changes
wb = openpyxl.load_workbook(excel_output_file)

# Iterate through all sheets in the workbook
for sheet in wb.sheetnames:
    ws = wb[sheet]
    
    for col in ws.columns:
        max_length = 0
        
        for cell in col:
            try: # Necessary to avoid error on empty cells
                  cell_value = str(cell.value)
                  if len(cell_value) > max_length:
                      max_length = len(cell_value)
                     
            except:
                  pass
         
            adjusted_width = (max_length + 2)
            # Set the column's width
            column_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width


    # Iterate through all cells in the sheet
    for row in ws.iter_rows():
        for cell in row:
            # Set alignment to center for both horizontal and vertical alignment
            cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

# Function to check if a value can be converted to a float
def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# Iterate through all sheets in the workbook
for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Specify the column index that you want to convert to numeric
    target_column_index = 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=target_column_index + 1, max_col=target_column_index + 1):
        for cell in row:
            cell_value = cell.value

            # Check if the cell contains text that can be converted to a float
            if cell_value is not None and isinstance(cell_value, str) and is_float(cell_value):
                cell.value = float(cell_value)

# Save the modified Excel file
wb.save(excel_output_file)
print(f"DataFrame has been written to {excel_output_file}")